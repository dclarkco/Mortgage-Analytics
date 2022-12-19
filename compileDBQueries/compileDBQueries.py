import pandas as pd
import openpyxl
import numpy as np
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
#import formulas
#import xlwings as xw
#import hashlib

#TODO

# Change exporter to use excel formulas, rather than bake them in using openpyxl or xlrwriter ~
# Allow for more flexibility in file type
# Modularize unused/one time features
# Add QR code generator

DBdir = "/Users/dclark/Documents/GitHub/Data/BatchProcessing"
# ^ All files in this directory will be processed into one file. be careful not to mix multiple search types together.

# batchProcess flags:
# out: export file or not
# excel: output as excel or csv
# compare_headers: Check for similar headers across all files
# compare_dates: check for missing date ranges
# split_dates: create two exported files depending on a specified daterange 
# analyze_rates: return analytics for data: rate ranges

def batchProcess(out, excel, compare_headers, compare_dates, analyze_rates, split_dates, check_address):

	fullDF = pd.DataFrame() #establish empty dataframe to fill with concatenated dataframes (from dataFrames list)
	dataFrames = [] #establish empty list of data frames to be concatenated later

	for filename in os.listdir(DBdir):  
		for root, dirs, files in os.walk(os.path.join(DBdir, filename)): # iterate through folders in BatchProcessing ->os.walk
			for file in files: 
				#print(root,dir,file)

				if "~" in file or ".DS_Store" in file: # skip null macOS files (excel prepends with ~ for temp or autosave files, which break this code)
					print("Null file skipped...")
					None

				#if excel and ".xlsx" in file:
				elif ".xlsx" in file:  # find excel files
					path = os.path.join(root, file) # create system path of file for use later
					print(path)

					exFile = openpyxl.load_workbook(path, data_only=True) # load excel file into a openpyxl workbook object
					print(type(exFile['do_not_modify_copy']), exFile['do_not_modify_copy']['A2'].value) # check file types

					val = exFile['do_not_modify_copy'].values # sheet 1 values saved as worksheet object
					ref = exFile['Lookups'].values # sheet 2 values saved as worksheet object
					col = next(val)[0:]	#get headers from column list (index 0)

					#tempDF = pd.read_excel(path, engine = 'openpyxl', parse_dates = ["mortgage_1_recording_date"]).sort_values("mortgage_1_recording_date")
					tempDF = pd.DataFrame(val, columns=col).drop(index=0) # for some reason, adding columns introduces a dupe row of columns, so this code drops that first row, keeping headers intact
					#refDF = pd.DataFrame(ref)

					print(tempDF['mortgage_1_recording_date'].head()) #check data integrity - prints the beginning of the dataframe to console

					print(file, "| Length:", len(tempDF), "| Start date:", tempDF["mortgage_1_recording_date"].iloc[0], "| End date:", tempDF["mortgage_1_recording_date"].iloc[-1])

					tempDF['mortgage_1_recording_date'] = pd.to_datetime(tempDF['mortgage_1_recording_date']) # convert dates to datetime object to avoid issues later on when matching these values against expected dates

					dataFrames.append(tempDF)#index_col = "mailing_full_street_address" # append 
					

				elif not excel and ".csv" in file: # code to parse CSV files instead of XLSX
					path = os.path.join(root, file)
					print(path)

					tempDF = pd.read_csv(path, low_memory = False, parse_dates = ["mortgage_1_recording_date"], infer_datetime_format=True).sort_values("mortgage_1_recording_date")
					#tempDF['mortgage_1_recording_date'] = pd.to_datetime(tempDF["mortgage_1_recording_date"], format = "%m-%d-%Y", utc=True)

					print(file, "| Length:", len(tempDF), "| Start date:", tempDF["mortgage_1_recording_date"].iloc[0], "| End date:", tempDF["mortgage_1_recording_date"].iloc[-1])
					
					dataFrames.append(tempDF)#index_col = "mailing_full_street_address"

				else:
					#print("Null file:", file)
					None
				
				#break
		# break

	fullDF = pd.concat(dataFrames, verify_integrity = False).sort_values("mortgage_1_recording_date").reindex() # concatenate dataframes into one, sort values by date, reset index

	print(type(fullDF.mortgage_1_recording_date)) # check type

	dfPre15 = fullDF[fullDF['mortgage_1_recording_date']< '02/01/2015'] # dates to be used for splitting dates, if needed. see split_dates
	dfPost15 = fullDF[fullDF['mortgage_1_recording_date']>= '02/01/2015']

	print(len(dfPre15), len(dfPost15)) #check length of each split df, datewise

	if compare_headers:
		comp_headers = compareHeaders(dataFrames) #see if any of the dataframes have differing headers
		if len(comp_headers) > 1:
			print("\nExcluded headers:", comp_headers)

	if compare_dates:
		compareDates(fullDF)

	if analyze_rates:
		analyzeRates(dfPre15, dfPost15)

	if out:
		if split_dates:

			if excel:
				dfPre15.to_excel(os.path.join(DBdir, "<2015_Complete.xlsx"), engine = 'XlsxWriter')
				dfPost15.to_excel(os.path.join(DBdir, ">2015_Complete.xlsx"), engine = 'XlsxWriter')
			elif not excel: 
				dfPre15.to_csv(os.path.join(DBdir, "<2015_Complete.csv"))
				dfPost15.to_csv(os.path.join(DBdir, ">2015_Complete.csv"))

		elif not split_dates:

			if excel:
				wb = Workbook()
				ws1 = wb.active
				ws1.title = "do_not_modify_copy" #create new, empty worksheet to export into a new xlsx file

				for r in dataframe_to_rows(fullDF, index=True, header=True): # create worksheet 1 from dataframe object
					ws1.append(r)

				ws2 = wb.create_sheet(title = "Lookups")
				ws2 = ref

				wb.save("/Users/dclark/Documents/GitHub/Data/BatchProcessing/Complete.xlsx")

				#fullDF.to_excel(os.path.join(DBdir, "Complete.xlsx", engine = 'XlsxWriter'))
			elif not excel: 
				fullDF.to_csv(os.path.join(DBdir, "Complete.csv"))
				
		#fullDF.to_excel(os.path.join(DBdir, "Complete.xlsx"))

	#print(fullDF["mortgage_1_recording_date"].describe(datetime_is_numeric=True))
	print("\nRates:\n", fullDF["current_interest_rate"].describe())

def compareHeaders(dataFrames): # compare dataset columns for discrepancies

	uniqueCols = []
	for i in range(len(dataFrames)):
		if i > 0:
			#print(len(dataFrames[i].columns))
			uniqueCols.append(dataFrames[i].columns.difference(dataFrames[i-1].columns).values)
	
	return uniqueCols[0]

def compareDates(fullDF):

	# TODO:
	# exclude holidays from comparison matrix

	allDates = pd.date_range(start = '2010-01-01', end = "2017-08-31", freq='B') # create index of 'business' dates between two dates (skips weekends)
	mortgageDates = pd.to_datetime(fullDF["mortgage_1_recording_date"]).unique() # export all unique dates from dataset
	missingDates = [] 

	print("DATE TIME DATA SIZE:", "CAPTURED DATES:", len(mortgageDates),"| ALL DATES: ", len(allDates))
	print("DATE TIME DATA TYPE:", "CAPTURED DATES:", type(mortgageDates),"| ALL DATES: ", type(allDates))
	print("DATE TIME DATA UNITS:", "CAPTURED DATES:", type(mortgageDates[1]),"| ALL DATES: ", type(allDates[1]))


	for date in allDates:
		if date in mortgageDates: # output all found or not found dates by comparing iteratively
			print(date, "Found...") # probablly really inefficient, but runtime isn't the goal here
		else:
			missingDates.append(date)
			print(date, "NOT Found...")

	print("\nMissing Days:", len(missingDates), "/", allDates) # output total amount of missing days. usually at least 20-100, as holidays are still counted.

	return missingDates

def analyzeRates(dfPre15, dfPost15): # in progress...
	#WARNING! SLOW...
	rateTopsVA = [3.88, 4.00, 4.13, 4.38, 4.63]
	rateTopsConv = [4.13, 4.25, 4.38, 4.63, 5.00]

	Mode = "Conv"

	if Mode == "VA":
		rateTops = rateTopsVA
	elif Mode == "Conv":
		rateTops = rateTopsConv
	print(rateTops)

	for rate in rateTops:
		print(rate)
		print((dfPre15['current_interest_rate'] <= rate).value_counts())
	uniques = dfPre15['current_interest_rate'].value_counts().sort_index()
	print(uniques,"-------------pre 15---------------")

	uniques.to_csv(os.path.join(DBdir, "<15_Uniques.csv"))

	for rate in rateTops:
		print(rate)
		print((dfPost15['current_interest_rate'] <= rate).value_counts())
	uniques = dfPost15['current_interest_rate'].value_counts().sort_index()
	print(uniques, "-------------post 15---------------")
	uniques.to_csv(os.path.join(DBdir, ">15_Uniques.csv"))

def checkAddress(addr):
	
	from usps import Address

	usps = USPSApi('544QED006465', test=True)
	
	validation = usps.validate_address(address)
	print(validation.result)




def generateQR(ref):

	import qrcode

	prepend = 'https://poweredbyreverse.com/results/GDM'

	link = "{}{}".format(prepend, ref) #formatting here
	
	# Creating an instance of QRCode class
	qr = qrcode.QRCode(version = 1,
	                   box_size = 10,
	                   border = 5)
	 
	# Adding data to the instance 'qr'
	qr.add_data(link)
	 
	qr.make(fit = True)
	img = qr.make_image(fill_color = 'black',
	                    back_color = 'white')
	 
	img.save('QR{}.png'.format(ref))

batchProcess(out = False, excel = False, compare_headers = False, compare_dates = False, split_dates = False, analyze_rates = False, check_address = True)

# main function:
# out: export file or not
# excel: output as excel or csv 
# compare_headers: Check for similar headers across all files
# compare_dates: check for missing date ranges
# split_dates: create two exported files depending on a specified daterange 
# analyze_rates: return analytics for data: rate ranges
